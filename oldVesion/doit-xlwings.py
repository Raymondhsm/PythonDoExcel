from os import path as osPath
from os import listdir,system
from datetime import datetime
from xlrd import open_workbook
from time import sleep
from threading import Thread
from openpyxl import load_workbook,Workbook

# define data class
class info:
    account = "undefined"
    name = "undefined"
    location = "undefined"
    salesAmount = 0.0
    profitRate = 0.0
    normal = True      # to mark the cell is nornal or not 


class refundInfo:
    platform = ""
    accountList = []
    locationList = []
    refundList = []
    length = 0

    def __init__(self, _platform):
        self.platform = _platform

    def add(self, _account, _location, _refund):
        self.accountList.append(_account)
        self.locationList.append(_location)
        self.refundList.append(_refund)
        self.length += 1

    def get(self, index):
        return self.accountList[index], self.locationList[index], self.refundList[index]

class error:
    errorType = "undefined"
    path = "undefined"
    message = "undefined"

    def __init__(self,_path,_message, _type = "Error"):
        self.errorType = _type
        self.path = _path
        self.message = _message

    def printError(self, _type):
        if _type == "Error":
            print("ERROR: " + self.message)
            print("PATH: " + self.path + "\n")
        
        if _type == "Warning":
            print("Warning: " + self.message)
            print("PATH: " + self.path + "\n")

        if _type == "NotFound":
            print("PATH: " + self.path)


def is_contains_chinese(strs):
    for _char in strs:
        if '\u4e00' <= _char <= '\u9fa5':
            return True
    return False

def getReportPath(path, _files, _onlyXlsx = True):
    _count = 0
    _pathList = []

    _isXls = False

    # print file list
    for _file in _files:
        # 去除文件夹，非.xlsx文件以及临时文件
        if (not osPath.isdir(path + '/' + _file)) and "~$" not in _file:
            if ".xlsx" in _file or (not _onlyXlsx and ".xls" in _file):
                _count += 1
                _pathList.append(path + '/' + _file)
                print(str(_count) + "、" + _file)

        
        # only at the onlyxlsx mode, will send the warning
        if _onlyXlsx and ".xls" in _file and ".xlsx" not in _file :
            _isXls = True

    # do not find file
    if _count == 0:
        print("do not find files")
        return None

    # warning find .xls file
    if _isXls:
        errorList.append(error("./", "We have found .xls file in report list", "Warning"))
        print("We have found .xls file, but the software can not read the .xls as report table!!!")
        print("Please convert it into .xlsx file, if you want to read it!!!")

    # input file number
    _index = 0
    while _index <= 0 or _index > _count:
        _input = input("input number to choose summary file:")
        if _input == "" or not _input.isdigit():
            continue
        
        _index = int(_input)
        if _index <= 0 or _index > _count:
            print("wrong number\n")
    
    return _pathList[_index - 1]


def findPlatformIndex(_reportTable, _platform):
    # store the merge cell's info
    mergeList = _reportTable.merged_cells
    mergeDict = {}
    for mergeCell in mergeList:
        mergeDict[mergeCell.min_row] = mergeCell.max_row - mergeCell.min_row

    # to find the index of platform
    isFindPlatform = False
    index = 1
    while index < _reportTable.max_row:
        _reportPlatform = _reportTable["C" + str(index)].value
        if _reportPlatform is None:
            index += 1
            continue
        _reportPlatform = _reportPlatform if not _reportPlatform.strip().isalpha() else _reportPlatform.lower()
        if _reportPlatform.strip() == _platform.strip().lower():
            isFindPlatform = True
            break
        else :
            if mergeDict.get(index) != None:
                index += mergeDict.get(index) + 1
            else :
                index +=1

    offset = mergeDict[index] + 1 if index in mergeDict else 1

    if isFindPlatform:
        return index, offset
    else:
        return -1, 0


def processRefundNotFound(_infoType, _platform, _refundAccount, _refundLocation, _refund):
    _startTime = datetime.now()
    for _row in range(2, notfoundTable.max_row + 1):
        _nfType = notfoundTable["A" + str(_row)].value
        _nfPlatform = notfoundTable["B" + str(_row)].value
        _nfAccount = notfoundTable["C" + str(_row)].value
        _nfLocation = notfoundTable["E" + str(_row)].value.split(' ')[1]

        if _infoType == _nfType and _platform == _nfPlatform and _refundAccount == _nfAccount and _refundLocation == _nfLocation:
            notfoundTable["H" + str(_row)] = _refund
            return

    _row = notfoundTable.max_row + 1
    notfoundTable["A" + str(_row)].value = _infoType
    notfoundTable["B" + str(_row)].value = _platform
    notfoundTable["C" + str(_row)].value = _refundAccount
    notfoundTable["E" + str(_row)].value = "unkown " + _refundLocation
    notfoundTable["I" + str(_row)].value = _refund

    _endTime = datetime.now()
    _interval = (_endTime-_startTime).seconds
    print("\rrefund insert，用时 " + str(_interval) + " 秒")
    errorList.append(error("REFUND ERROR!", "Can not match the refund account and location!!"))


def getRefundInfo(_refundTable):
    _refundInfoList = []

    # use the merge cell to locate the useful cell
    for merge in _refundTable.merged_cells:
        rs, re, cs, ce = merge

        # ignore some merge cell
        if re-rs != 1 and ce-cs != 2:
            continue

        # ignore when it has no info
        if _refundTable.cell_value(re,cs) == "":
            continue
        
        # read the platform
        _platform = _refundTable.cell_value(rs,cs).strip().lower()
        if is_contains_chinese(_platform):
            errorList.append(error("REFUND ERROR! Type = " + _refundTable.name + "\tPlatform: " + _platform, "We can not identify the chinese as platform name"))
            continue
        else:
            _refundInfoInstance = refundInfo(_platform)

        _row = re + 1 if _refundTable.cell_value(re,cs) == "账号" else re
        while True:
            _acclo = _refundTable.cell_value(_row,cs)

            if _acclo == "":
                break
            else:
                # split the account and location
                if '(' in _acclo:
                    _accloList = _acclo.split("(",1)
                    _account = _accloList[0].strip()
                    _location = _accloList[1].split(')',1)[0].strip()
                elif '（' in _acclo:
                    _accloList = _acclo.split("（",1)
                    _account = _accloList[0].strip()
                    _location = _accloList[1].split('）',1)[0].strip()
                else:
                    _acclo = _acclo.strip()
                    _accloList = _acclo.split(" ",1)
                    _account = _accloList[0].strip()

                    # set the default value
                    if len(_accloList) == 1:
                        _location = "CN"
                    else:
                        _location = _accloList[1].strip()
                
                # read the refund
                _refund = _refundTable.cell_value(_row,cs+1)
                
                _refundInfoInstance.add(_account, _location, _refund)
            
            # add 1 to row
            _row += 1
        
        # add to the list
        _refundInfoList.append(_refundInfoInstance)

    return _refundInfoList

def setRefundInfo(_infoType,  _refundInfoList):
    # if can not find correct type, return
    if _infoType not in summary.sheetnames:
        errorList.append(error("REFUND ERROR","can not find correct sheet: " + _infoType))
        return 
    _refundTable = summary[_infoType]

    for _refundInstance in _refundInfoList:
        _index, _offset = findPlatformIndex(_refundTable, _refundInstance.platform)
        
        # if can not find platform, send error
        if _index == -1:
            errorList.append(error("REFUND ERROR","can not find correct platform: " + _refundInstance.platform))
            continue
        
        # create the dictionary 
        accountDict = {}
        for _row in range(_index, _index + _offset):
            _account = _refundTable["D" + str(_row)].value.strip()
            _locationList = _refundTable["F" + str(_row)].value.split(' ')
            _location = _locationList[1].strip() if len(_locationList) > 1 else _locationList[0].strip()

            # add to the dict
            accountDict[_account + '_' + _location] = _row

        # process the refund
        for _refundIndex in range(0, _refundInstance.length):
            _refundAccount, _refundLocation, _refund = _refundInstance.get(_refundIndex)

            if _refundAccount + '_' + _refundLocation in accountDict:
                _refundTable["I" + str(accountDict[_refundAccount + '_' + _refundLocation])].value = _refund
            else:
                processRefundNotFound(_infoType, _refundInstance.platform, _refundAccount, _refundLocation, _refund)

    return

def processRefundInfo(_filePath):
    _refundBook = open_workbook(_filePath,formatting_info=True)
    _refundSheets = _refundBook.sheets()


    for _refundsheet in _refundSheets:
        _infoType = _refundsheet.name

        print("正在处理 " + _refundsheet.name + " 退款金额...\r")

        _refundInfoList = getRefundInfo(_refundsheet)
        setRefundInfo(_infoType, _refundInfoList)


    return


def getInfoByXlrd(filePath):
    # open the file
    data = open_workbook(filePath)

    # to store this file data which _platform and type is
    _platform = data.sheet_by_index(0).cell_value(1,0).split('/')[0]
    _infoType = data.sheet_by_index(0).cell_value(1,1).split('/',2)[1]

    # use info list to temporary store data 
    _infoList = []

    # for each table to process data
    sheets = data.sheets()
    for table in sheets:
        # ignore the sheet of "原始"
        if table.name == "原始":
            continue

        _row = 1
        nrow = table.nrows

        # for each module to find out useful info
        while _row < nrow:
            # if find the none row, add 1 to row for finding the next
            if table.cell(_row,1).value == "":
                _row += 1
                continue

            # if account is none, the wo think the line is bad
            row_3List = table.cell(_row,2).value.split('/')
            if row_3List[0] == "" or is_contains_chinese(row_3List[0]):
                _row += 1
                continue

            # create the module info instance
            _infoInstance = info()

            # to process the account/location cell
            _infoInstance.account = row_3List[0]
            _infoInstance.location = row_3List[1] if not is_contains_chinese(row_3List[1]) else row_3List[1][0:-1]

            # if cell has no name, then wo think this line is bad 
            _nameInTable = table.cell(_row,1).value.split('/',2)
            _infoInstance.name = "/" if _nameInTable[0] == "" else _nameInTable[0]
            if _infoType == "类": _infoType = _nameInTable[1]
    
            # to judge this module is normal or not
            row_4 = table.cell(_row,3).value
            if row_4 == "":
                _infoInstance.normal = False
                _infoInstance.salesAmount = table.cell(_row,4).value
            else :
                _infoInstance.salesAmount = row_4
                _infoInstance.profitRate = table.cell(_row+1,4).value

            # append the module instance into the list
            _infoList.append(_infoInstance)

            # add 3 to variable _row to move to next module 
            _row += 2

    data.release_resources()
    if _infoType == "C类" : _infoType += ("打底裤")
    return _platform, _infoType, _infoList


def openReport(_reportPath):
    print("正在读取汇总文件\r")
    try:
        _summary = load_workbook(_reportPath)
    except Exception as e:
        print("读取汇总文件发生错误，可能汇总文件已被打开，请关闭文件后重试！！")
        print("Unexpected Error: {}".format(e))
    else:    
        print("汇总文件读取成功：" +  _reportPath + "\n")

    return _summary

def initNotFoundTable():
    notfoundTable["A1"] = "类型"
    notfoundTable["B1"] = "渠道"
    notfoundTable["C1"] = "账号"
    notfoundTable["D1"] = "姓名"
    notfoundTable["E1"] = "姓名/仓"
    notfoundTable["F1"] = "销售额"
    notfoundTable["G1"] = "利润率"
    notfoundTable["H1"] = "退款金额"
    notfoundTable["I1"] = "毛利"

def save(_fileName):
    print("\n正在保存文件")
    _fileName = "output" if _fileName == "" else _fileName

    try:
        summary.save(path + "/" + _fileName + ".xlsx")
        notfound.save(path + "/notfound.xlsx")

    except Exception as e:
        errorList.append(error(path + '/' + _fileName + ".xlsx", "Unexpected Error: {}".format(e)))
        print("文件保存失败，请检查 " + _fileName + ".xlsx 或 notfound.xlsx 是否在打开状态。请关闭文件后重试！！！")
        print("Unexpected Error: {}".format(e))

    else:
        print("保存文件成功，路径：" + path + "/" + _fileName + ".xlsx") 

# --------------------------------------------------------------------------------------------- #
# main function #
# --------------------------------------------------------------------------------------------- #


#文件夹目录
path = "../data/"
#得到文件夹下的所有文件名称
files= listdir(path) 

# mark the message throught the process
errorList = []
correctCount = 0
failCount = 0

# open report to write down
reportPath = getReportPath(path, files)
refundPath = getReportPath(path, files, _onlyXlsx = False)
notfound = Workbook()
notfoundTable = notfound.active
initNotFoundTable()

if reportPath is not None:
    # open the report
    summary = openReport(reportPath)
    processRefundInfo(refundPath)

# init notfound table


save("output")
    
# pause the os in case disappear
system("pause")

    
            
                

               