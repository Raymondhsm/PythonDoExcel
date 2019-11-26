from openpyxl import load_workbook
from copy import copy

insert = load_workbook("../data/insert.xlsx")
table = insert[insert.sheetnames[0]]

outmerges = table.merged_cells
# for merge in merges:
#     print(merge.min_row)
#     print(merge.max_col)


def copy_cell(source_cell, target_cell):
    target_cell.data_type = source_cell.data_type
    target_cell.value = source_cell.value
    target_cell.fill = copy(source_cell.fill)
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)
 
 
def insert_rows(table, row, attach_direction = None,amount = 1):
    '''
    :Description: 重新封装插入一行
    \n:param: 
        table   需要插入的表格sheet
        row     插入的行号
        amount  插入的行数
        attach_direction    None为不合并，True为附着在上方的合并单元格，False为附着在下方的合并单元格
    \n:return: none
    '''

    # 获取合并的单元格
    merges = table.merged_cells
    mergesCpy = merges.__copy__()

    # 拆分行号大于插入值的合并单元格，方便处理
    for merge in mergesCpy:
        if merge.max_row >= row:
            table.unmerge_cells(merge.coord)

    # 插入新的行
    table.insert_rows(row, amount)

    # 重新合并单元格
    for merge in mergesCpy:
        if attach_direction is None:
            min_row = merge.min_row + amount if merge.min_row >= row else merge.min_row
            max_row = merge.max_row + amount if merge.max_row >= row else merge.max_row

        elif attach_direction:
            min_row = merge.min_row + amount if merge.min_row >= row else merge.min_row
            max_row = merge.max_row + amount if merge.max_row + 1 >= row else merge.max_row
            
        elif not attach_direction:
            # 复制下方单元格
            if merge.min_row == row:
                target_cell = table.cell(row, merge.min_col)
                source_cell = table.cell(row + amount, merge.min_col)
                copy_cell(source_cell, target_cell)

            min_row = merge.min_row + amount if merge.min_row > row else merge.min_row
            max_row = merge.max_row + amount if merge.max_row >= row else merge.max_row
        
        table.merge_cells(None, min_row, merge.min_col, max_row, merge.max_col)


def insert_cols(table, col, attach_direction = None, amount = 1):
    '''
    :Description: 重新封装插入一列
    \n:param: 
        table   需要插入的表格sheet
        col     插入的列号
        amount  插入的列数
        attach_direction    None为不合并，True为附着在左方的合并单元格，False为附着在右方的合并单元格
    \n:return: none
    '''

    # 获取合并的单元格
    merges = table.merged_cells
    mergesCpy = merges.__copy__()

    # 拆分行号大于插入值的合并单元格，方便处理
    for merge in mergesCpy:
        if merge.max_col >= col:
            table.unmerge_cells(merge.coord)

    # 插入新的行
    table.insert_cols(col, amount)

    # 重新合并单元格
    for merge in mergesCpy:
        if attach_direction is None:
            min_col = merge.min_col + amount if merge.min_col >= col else merge.min_col
            max_col = merge.max_col + amount if merge.max_col >= col else merge.max_col

        elif attach_direction:
            min_col = merge.min_col + amount if merge.min_col >= col else merge.min_col
            max_col = merge.max_col + amount if merge.max_col + 1 >= col else merge.max_col
            
        elif not attach_direction:
            # 复制下方单元格
            if merge.min_col == col:
                target_cell = table.cell(merge.min_row, col)
                source_cell = table.cell(merge.min_row, col + amount)
                copy_cell(source_cell, target_cell)

            min_col = merge.min_col + amount if merge.min_col > col else merge.min_col
            max_col = merge.max_col + amount if merge.max_col >= col else merge.max_col
        
        table.merge_cells(None, merge.min_row, min_col, merge.max_row, max_col)


table['A5'] = 10

insert_cols(table,4,False)
# insert_rows(table,12,1,True)
# insert_rows(table,12,1,False)

insert.save("../data/ooo.xlsx")