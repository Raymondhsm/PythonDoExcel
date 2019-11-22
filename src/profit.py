from error import processException, ErrorList, NotFound
from xlrd import open_workbook
from openpyxl import load_workbook,Workbook

class profitInfo:
    
    def __init__(self, pid, amount):
        self.id = pid
        self.amount = amount


class Profit:

    def __init__(self, originPath, updatePath):
        self.updatePath = updatePath
        try:
            originBook = open_workbook(originPath)
            self.originTable = originBook.sheet_by_index(0)

            self.updateBook = load_workbook(updatePath)
            self.updateTable = self.updateBook[self.updateBook.sheetnames[0]]

            self.pidDict = self.__setupDict()
        except:
            processException()

    
    def processProfitUpdate(self):
        print("正在读取文件...")
        profitList = self.__readProfit()
        print("正在写入文件...")
        self.__writeProfit(profitList)

    def save(self, outPath = ""):
        try:
            print("正在保存文件...")
            if outPath == "":
                self.updateBook.save(self.updatePath)
            else:
                self.updateBook.save(outPath)
            print("保存文件成功！！")
        except:
            processException()


    def __readProfit(self):
        try:
            row = self.originTable.nrows

            profitList = []

            for index in range(1, row):
                pid = self.originTable.cell_value(index, 0)
                amount = self.originTable.cell_value(index, 1)
                profitList.append(profitInfo(pid, amount))

            return profitList
        except:
            processException()


    def __setupDict(self):
        try:
            row = self.updateTable.max_row

            pidDict = {}

            for index in range(1, row):
                pidDict[self.updateTable["A" + str(index)].value] = index

            return pidDict
        except:
            processException()


    def __writeProfit(self, profitList):
        try:
            for profitInstance in profitList:
                if profitInstance.id in self.pidDict:
                    index = self.pidDict[profitInstance.id]
                    self.updateTable["C" + str(index)] = profitInstance.amount
                else:
                    ErrorList.addError(NotFound("UPDATE ERROR!!! ID = " + profitInstance.id, ''))
        except:
            processException()

    
    



