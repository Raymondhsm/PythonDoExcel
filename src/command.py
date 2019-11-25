from error import ErrorList, Error, Warning, processException
import sys
from xlrd import open_workbook
from openpyxl import load_workbook,Workbook
import profit
import saleAmount
import refund
import utils
import notfound
from log import Logger


def doCommand():
    # forever loop
    while True:
        # split the command and remove the none
        cmdLine = input(">>-->")
        Logger.addLog("COMMAND: " + cmdLine)
        if cmdLine == "" : continue

        cmdList = cmdLine.strip().split(" ")
        for cmdIteration in cmdList:
            if cmdIteration == '' : cmdList.remove(cmdIteration)

        # we have no different cmd now, so ignore it
        # cmd = cmdList[0]

        index = 1
        result = {}
        while index < len(cmdList):
            # if it is the path Argument, then save it
            if cmdList[index] in pathArguDict:
                index += 1
                if index >= len(cmdList):
                    Logger.addLog("DATA ERROR: " + cmdList[index-1])
                    print("data none!!" + cmdList[index-1])
                    break
                else:
                    Logger.addLog("COMMAND: {}, DATA: {}".format(cmdList[index-1], cmdList[index]))
                    result[pathArguDict[cmdList[index-1]]] = cmdList[index]
                    index += 1

            elif cmdList[index] in systemArguDict:
                Logger.addLog("COMMAND: {}".format(cmdList[index]))
                systemArguDict[cmdList[index]]()
                break

            else:
                Logger.addLog("COMMAND ERROR:" + cmdList[index])
                print("command error! " + cmdList[index])
                break
        
        # 处理金额和退款
        if "salePath" in result and "summaryPath" in result and "refundPath" in result:
            Logger.addLog("process SA with RF")
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process salesAmount
            SA = saleAmount.SaleAmount(summary, notfoundTable)
            SA.processDir(result["salePath"])

            # process refund
            RF = refund.Refund(summary,notfoundTable)
            RF.processRefundInfo(result["refundPath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            summary.save(savePath)
            Logger.addPrefabLog(Logger.LOG_TYPE_SAVE,savePath)
            NF.save()

        # 处理金额
        elif "salePath" in result and "summaryPath" in result and "refundPath" not in result:
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process salesAmount
            SA = saleAmount.SaleAmount(summary, notfoundTable)
            SA.processDir(result["salePath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            Logger.addPrefabLog(Logger.LOG_TYPE_SAVE,savePath)
            summary.save(savePath)

        # 处理退款
        elif "summaryPath" in result and "refundPath" in result:
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process refund
            RF = refund.Refund(summary,notfoundTable)
            RF.processRefundInfo(result["refundPath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            summary.save(savePath)
            Logger.addPrefabLog(Logger.LOG_TYPE_SAVE,savePath)
            

        # 更新成本
        elif "originPath" in result and "updatePath" in result:
            PF = profit.Profit(result["originPath"],result["updatePath"])
            PF.processProfitUpdate()

            savePath = result["savePath"] if "savePath" in result else ""
            PF.save(savePath)

        # print
        if len(result) != 0 : ErrorList.printErrorList()


def printVersion():
    print("6.0 By XiaoMing \n")

def printHelp():
    for pa in pathArguDict.keys():
        print(("%-10.5s" % pa) + pathArguDict[pa])

    print(("%-10.5s" % "-v") + "查看版本")
    print(("%-10.5s" % "-help") + "查看帮助")
    print(("%-10.5s" % "-exit") + "退出程序")

    print("for example: do -o filePath -u filePath")



cmdDict = {
    # "do" : 
}

pathArguDict = {
    "-o" : "originPath", 
    "-u" : "updatePath",
    "-sa" : "salePath",
    "-r" : "refundPath",
    "-su" : "summaryPath",
    "-out" : "savePath",
    "-n" : "notfoundPath"
}

systemArguDict = {
    "-v" : printVersion,
    "-help" : printHelp,
    "-exit" : sys.exit
}


        

    

