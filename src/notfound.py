from error import processException
from log import Logger
from openpyxl import Workbook,load_workbook

class NotFound:

    def __init__(self, path = ""):
        Logger.addLog("CREAT NotFound: " + path)
        if path == "":
            self.notfoundBook = Workbook()
            self.path = "notfound.xlsx"
        else:
            try:
                self.notfoundBook = load_workbook(path)
                self.path = path
            except:
                processException()


    def getNotfoundTable(self):
        Logger.addLog("GET notfoundTable!!")
        try:
            if len(self.notfoundBook.sheetnames) == 0:
                notfoundTable = self.notfoundBook.active
                self.__initNotFoundTable(notfoundTable)
            else:
                notfoundTable = self.notfoundBook[self.notfoundBook.sheetnames[0]]
            
            return notfoundTable
        except:
            processException()


    def save(self, notfoundOutPath = ""):
        Logger.addPrefabLog(Logger.LOG_TYPE_SAVE, notfoundOutPath)
        if notfoundOutPath == "":
            self.notfoundBook.save(self.path)
        else:
            self.notfoundBook.save(notfoundOutPath)


    def __initNotFoundTable(self, notfoundTable):
        notfoundTable["A1"] = "类型"
        notfoundTable["B1"] = "渠道"
        notfoundTable["C1"] = "账号"
        notfoundTable["D1"] = "姓名"
        notfoundTable["E1"] = "姓名/仓"
        notfoundTable["F1"] = "销售额"
        notfoundTable["G1"] = "利润率"
        notfoundTable["H1"] = "退款金额"
        notfoundTable["I1"] = "毛利"
