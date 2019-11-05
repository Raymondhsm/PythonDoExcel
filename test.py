import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

# open the file
# xlrd.Book.encoding = "utf8"
# data = xlrd.open_workbook("A.xlsx")
# summary = xlrd.open_workbook("B.xls",formatting_info=True)

# table = summary.sheet_by_index(0)

# print(table.merged_cells)

str = "Hel  lo"
print(str.split(" ",1))

# open report to write down
# summary = xlrd.open_workbook("../data/refund.xls",formatting_info=True)
# newbook = Workbook()

# find the infotype table
# reportTable = summary.sheet_by_index(0)
# print(reportTable.name)

# for merge in reportTable.merged_cells:
#         rs, re, cs, ce = merge
#         print(1)
#         print(merge)

# newsheet = newbook.create_sheet("test")
# newsheet = summary.copy_worksheet(reportTable)

# # find the index of platform
# mergeList = reportTable.merged_cells

# print(mergeList)
# fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)
# reportTable['G3'].value = "safdasfa"
# reportTable['G3'].font = fontStyle
# summary.save("output.xlsx")

# newbook.save("output.xlsx")
# summary.close()
# newbook.close()